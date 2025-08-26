#!/usr/bin/env python3
"""
Generate Excel Report for Corewell Health Open Payments Analysis
DA-175: Multi-sheet Excel workbook with comprehensive analysis results

This script consolidates all analysis outputs into a structured Excel workbook
with formatting, charts, and executive summary.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
import glob
import json

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Define paths
PROJECT_DIR = Path("/home/incent/conflixis-data-projects/projects/175-corewell-health-analysis")
DATA_DIR = PROJECT_DIR / "op-payment-report" / "data" / "processed"
OUTPUT_DIR = PROJECT_DIR / "op-payment-report" / "reports"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Define styles
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0C343A", end_color="0C343A", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4C94ED", end_color="4C94ED", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def load_latest_data():
    """Load the most recent analysis outputs"""
    data = {}
    
    # Define file patterns
    patterns = {
        'op_overall': 'op_overall_metrics_*.csv',
        'op_yearly': 'op_yearly_trends_*.csv',
        'op_categories': 'op_payment_categories_*.csv',
        'op_manufacturers': 'op_top_manufacturers_*.csv',
        'op_tiers': 'op_payment_tiers_*.csv',
        'op_consecutive': 'op_consecutive_years_*.csv',
        'rx_overall': 'rx_overall_metrics_*.csv',
        'rx_yearly': 'rx_yearly_trends_*.csv',
        'rx_top_drugs': 'rx_top_drugs_*.csv',
        'rx_high_cost': 'rx_high_cost_drugs_*.csv',
        'rx_specialty': 'rx_by_specialty_*.csv',
        'rx_np_pa': 'rx_np_pa_analysis_*.csv',
        'corr_drug': 'correlation_by_drug_*.csv',
        'corr_provider': 'correlation_by_provider_type_*.csv',
        'corr_payment_tier': 'correlation_by_payment_tier_*.csv',
        'corr_consecutive': 'correlation_consecutive_years_*.csv'
    }
    
    for key, pattern in patterns.items():
        files = list(DATA_DIR.glob(pattern))
        if files:
            latest_file = max(files, key=lambda x: x.stat().st_mtime)
            logger.info(f"Loading {key}: {latest_file.name}")
            data[key] = pd.read_csv(latest_file)
        else:
            logger.warning(f"No files found for pattern: {pattern}")
            
    return data

def create_executive_summary(wb, data):
    """Create executive summary sheet"""
    ws = wb.create_sheet("Executive Summary", 0)
    
    # Title
    ws['A1'] = "Corewell Health Open Payments Analysis - Executive Summary"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    # Key metrics section
    row = 3
    ws[f'A{row}'] = "KEY METRICS (2020-2024)"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].fill = HEADER_FILL
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    metrics = []
    
    # Open Payments metrics
    if 'op_overall' in data:
        op = data['op_overall'].iloc[0]
        metrics.extend([
            ("Providers Receiving Payments", f"{op['unique_providers']:,.0f}", "73.5% of all Corewell providers"),
            ("Total Payment Transactions", f"{op['total_transactions']:,.0f}", ""),
            ("Total Payments Received", f"${op['total_payments']:,.2f}", ""),
            ("Average Payment Amount", f"${op['avg_payment']:,.2f}", ""),
            ("Maximum Single Payment", f"${op['max_payment']:,.2f}", "")
        ])
    
    # Prescription metrics
    if 'rx_overall' in data:
        rx = data['rx_overall'].iloc[0]
        metrics.extend([
            ("", "", ""),  # Blank row
            ("Total Prescribers", f"{rx['unique_prescribers']:,.0f}", "92.6% of all Corewell providers"),
            ("Total Prescriptions", f"{rx['total_prescriptions']:,.0f}", ""),
            ("Total Prescription Value", f"${rx['total_payments']:,.2f}", ""),
            ("Unique Drugs Prescribed", f"{rx['unique_drugs']:,.0f}", "")
        ])
    
    for metric, value, note in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        ws[f'C{row}'] = note
        if metric:  # Only format non-blank rows
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(size=12)
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
        row += 1
    
    # Critical findings section
    row += 2
    ws[f'A{row}'] = "CRITICAL FINDINGS"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    findings = [
        "1. EXTREME PAYMENT INFLUENCE:",
        "   • Krystexxa: Providers with payments prescribe 426x more than those without",
        "   • Enbrel: 218x increase with payments",
        "   • Trelegy: 115x increase with payments",
        "",
        "2. PA/NP VULNERABILITY:",
        "   • Physician Assistants: 407.6% increase in prescribing when receiving payments",
        "   • Nurse Practitioners: 280.2% increase with payments",
        "",
        "3. SUSTAINED RELATIONSHIPS:",
        "   • 2,342 providers (22.5%) received payments EVERY year from 2020-2024",
        "",
        "4. MINIMAL THRESHOLD FOR INFLUENCE:",
        "   • Even $1-100 payments show 23,218x ROI",
        "   • Payment influence starts at smallest payment levels"
    ]
    
    for finding in findings:
        ws[f'A{row}'] = finding
        if finding.startswith(("1.", "2.", "3.", "4.")):
            ws[f'A{row}'].font = Font(bold=True, color="FF0000")
            ws[f'A{row}'].fill = HIGHLIGHT_FILL
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    # Compliance risks
    row += 2
    ws[f'A{row}'] = "COMPLIANCE RISK INDICATORS"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    risks = [
        "• Anti-Kickback Statute: Extreme correlations suggest potential violations",
        "• False Claims Act: Payment-influenced prescribing may constitute false claims",
        "• Stark Law: Financial relationships may violate self-referral prohibitions",
        "• Transparency: 73.5% of providers receiving payments raises questions"
    ]
    
    for risk in risks:
        ws[f'A{row}'] = risk
        ws[f'A{row}'].font = Font(color="FF6600")
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    
    return ws

def create_payment_analysis_sheet(wb, data):
    """Create Open Payments analysis sheet"""
    ws = wb.create_sheet("Payment Analysis")
    
    row = 1
    ws[f'A{row}'] = "Open Payments Analysis"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:G{row}')
    
    # Yearly trends
    if 'op_yearly' in data:
        row += 3
        ws[f'A{row}'] = "Yearly Payment Trends"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws.merge_cells(f'A{row}:G{row}')
        
        row += 1
        df = data['op_yearly']
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=row, column=col_idx, value=col).font = HEADER_FONT
            ws.cell(row=row, column=col_idx).fill = HEADER_FILL
            
        for _, data_row in df.iterrows():
            row += 1
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=value)
    
    # Top manufacturers
    if 'op_manufacturers' in data:
        row += 3
        ws[f'A{row}'] = "Top 20 Manufacturers by Payments"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws.merge_cells(f'A{row}:G{row}')
        
        row += 1
        df = data['op_manufacturers'].head(20)
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=row, column=col_idx, value=col).font = HEADER_FONT
            ws.cell(row=row, column=col_idx).fill = HEADER_FILL
            
        for _, data_row in df.iterrows():
            row += 1
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=value)
    
    # Payment categories
    if 'op_categories' in data:
        row += 3
        ws[f'A{row}'] = "Payment Categories"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws.merge_cells(f'A{row}:G{row}')
        
        row += 1
        df = data['op_categories'].head(15)
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=row, column=col_idx, value=col).font = HEADER_FONT
            ws.cell(row=row, column=col_idx).fill = HEADER_FILL
            
        for _, data_row in df.iterrows():
            row += 1
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    return ws

def create_prescription_analysis_sheet(wb, data):
    """Create prescription patterns analysis sheet"""
    ws = wb.create_sheet("Prescription Patterns")
    
    row = 1
    ws[f'A{row}'] = "Prescription Pattern Analysis"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:H{row}')
    
    # Top drugs
    if 'rx_top_drugs' in data:
        row += 3
        ws[f'A{row}'] = "Top 30 Drugs by Total Payments"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        df = data['rx_top_drugs'].head(30)
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=row, column=col_idx, value=col).font = HEADER_FONT
            ws.cell(row=row, column=col_idx).fill = HEADER_FILL
            
        for _, data_row in df.iterrows():
            row += 1
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=value)
    
    # High-cost targeted drugs
    if 'rx_high_cost' in data:
        row += 3
        ws[f'A{row}'] = "Key High-Cost Drugs Analysis"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        df = data['rx_high_cost']
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=row, column=col_idx, value=col).font = HEADER_FONT
            ws.cell(row=row, column=col_idx).fill = HEADER_FILL
            
        for _, data_row in df.iterrows():
            row += 1
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=value)
    
    # Specialty prescribing
    if 'rx_specialty' in data:
        row += 3
        ws[f'A{row}'] = "Prescribing by Specialty"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        df = data['rx_specialty'].head(20)
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=row, column=col_idx, value=col).font = HEADER_FONT
            ws.cell(row=row, column=col_idx).fill = HEADER_FILL
            
        for _, data_row in df.iterrows():
            row += 1
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    return ws

def create_correlation_analysis_sheet(wb, data):
    """Create payment-prescription correlation analysis sheet"""
    ws = wb.create_sheet("Payment Correlations")
    
    row = 1
    ws[f'A{row}'] = "Payment-Prescription Correlation Analysis"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:J{row}')
    
    # Drug correlations
    if 'corr_drug' in data:
        row += 3
        ws[f'A{row}'] = "Drug-Level Payment Influence (Paid vs Unpaid Providers)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws.merge_cells(f'A{row}:J{row}')
        
        row += 1
        df = data['corr_drug'].head(30)
        
        # Highlight extreme correlations
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=row, column=col_idx, value=col).font = HEADER_FONT
            ws.cell(row=row, column=col_idx).fill = HEADER_FILL
            
        for _, data_row in df.iterrows():
            row += 1
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                # Highlight extreme ratios
                if col_idx == df.columns.get_loc('prescribing_ratio') + 1 if 'prescribing_ratio' in df.columns else -1:
                    if isinstance(value, (int, float)) and value > 100:
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        cell.font = Font(bold=True, color="FF0000")
    
    # Provider type correlations
    if 'corr_provider' in data:
        row += 3
        ws[f'A{row}'] = "Provider Type Payment Influence"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws.merge_cells(f'A{row}:J{row}')
        
        row += 1
        df = data['corr_provider']
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=row, column=col_idx, value=col).font = HEADER_FONT
            ws.cell(row=row, column=col_idx).fill = HEADER_FILL
            
        for _, data_row in df.iterrows():
            row += 1
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                # Highlight PA/NP rows
                if col_idx == 1 and value in ['PA', 'NP']:
                    for c in range(1, len(data_row) + 1):
                        ws.cell(row=row, column=c).fill = HIGHLIGHT_FILL
    
    # Payment tier analysis
    if 'corr_payment_tier' in data:
        row += 3
        ws[f'A{row}'] = "Payment Tier Influence Analysis"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws.merge_cells(f'A{row}:J{row}')
        
        row += 1
        df = data['corr_payment_tier']
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=row, column=col_idx, value=col).font = HEADER_FONT
            ws.cell(row=row, column=col_idx).fill = HEADER_FILL
            
        for _, data_row in df.iterrows():
            row += 1
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=value)
    
    # Consecutive years
    if 'corr_consecutive' in data:
        row += 3
        ws[f'A{row}'] = "Consecutive Year Payment Impact"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws.merge_cells(f'A{row}:J{row}')
        
        row += 1
        df = data['corr_consecutive']
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=row, column=col_idx, value=col).font = HEADER_FONT
            ws.cell(row=row, column=col_idx).fill = HEADER_FILL
            
        for _, data_row in df.iterrows():
            row += 1
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    return ws

def create_recommendations_sheet(wb):
    """Create recommendations and action items sheet"""
    ws = wb.create_sheet("Recommendations")
    
    row = 1
    ws[f'A{row}'] = "Compliance Recommendations & Action Items"
    ws[f'A{row}'].font = Font(bold=True, size=16)
    ws.merge_cells(f'A{row}:D{row}')
    
    recommendations = [
        ("IMMEDIATE ACTIONS", "Priority", "Timeline", [
            ("Implement real-time monitoring for providers with >$1,000 annual payments", "Critical", "30 days"),
            ("Flag prescribing patterns deviating >50% from peer averages", "Critical", "30 days"),
            ("Develop specialized oversight program for PA/NP providers", "Critical", "45 days"),
            ("Require additional approval for high-cost drug prescriptions by PAs/NPs", "High", "60 days"),
            ("Mandatory disclosure of all industry interactions >$100", "High", "30 days")
        ]),
        ("TRANSPARENCY INITIATIVES", "Priority", "Timeline", [
            ("Public disclosure of all payments >$100", "High", "60 days"),
            ("Quarterly reporting to compliance committee", "Medium", "90 days"),
            ("Patient notification of provider industry relationships", "Medium", "120 days"),
            ("Annual transparency report publication", "Medium", "180 days")
        ]),
        ("EDUCATION PROGRAMS", "Priority", "Timeline", [
            ("Mandatory annual training on appropriate industry interactions", "High", "90 days"),
            ("Case studies showing payment influence on prescribing", "Medium", "90 days"),
            ("Clear guidelines on acceptable vs. problematic relationships", "High", "60 days"),
            ("Specialty-specific training for high-risk areas", "Medium", "120 days")
        ]),
        ("AUDIT & MONITORING", "Priority", "Timeline", [
            ("Quarterly audits of high-risk providers", "Critical", "Immediate"),
            ("Review of prescribing patterns for promoted drugs", "High", "30 days"),
            ("Investigation of outlier prescribing behaviors", "High", "45 days"),
            ("Third-party payment channel review", "Medium", "90 days")
        ]),
        ("POLICY DEVELOPMENT", "Priority", "Timeline", [
            ("Establish maximum annual payment thresholds", "High", "90 days"),
            ("Prohibit consecutive year payments from same manufacturer", "Medium", "120 days"),
            ("Restrict high-value meals and entertainment", "High", "60 days"),
            ("Develop conflict of interest policy", "Critical", "45 days")
        ])
    ]
    
    row = 3
    for category, priority_header, timeline_header, items in recommendations:
        # Category header
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Subheaders
        ws[f'A{row}'] = "Action Item"
        ws[f'B{row}'] = priority_header
        ws[f'C{row}'] = timeline_header
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws[f'B{row}'].fill = SUBHEADER_FILL
        ws[f'C{row}'].fill = SUBHEADER_FILL
        row += 1
        
        # Items
        for item, priority, timeline in items:
            ws[f'A{row}'] = item
            ws[f'B{row}'] = priority
            ws[f'C{row}'] = timeline
            
            # Color-code priority
            if priority == "Critical":
                ws[f'B{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
            elif priority == "High":
                ws[f'B{row}'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
            
            row += 1
        
        row += 1  # Blank row between categories
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    
    return ws

def main():
    """Main execution function"""
    logger.info("=" * 80)
    logger.info("GENERATING EXCEL REPORT FOR COREWELL HEALTH ANALYSIS")
    logger.info("=" * 80)
    
    try:
        # Load all data
        logger.info("Loading analysis data...")
        data = load_latest_data()
        
        if not data:
            logger.error("No data files found. Please run analysis scripts first.")
            return
        
        # Create workbook
        logger.info("Creating Excel workbook...")
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        logger.info("Creating Executive Summary...")
        create_executive_summary(wb, data)
        
        logger.info("Creating Payment Analysis sheet...")
        create_payment_analysis_sheet(wb, data)
        
        logger.info("Creating Prescription Patterns sheet...")
        create_prescription_analysis_sheet(wb, data)
        
        logger.info("Creating Correlation Analysis sheet...")
        create_correlation_analysis_sheet(wb, data)
        
        logger.info("Creating Recommendations sheet...")
        create_recommendations_sheet(wb)
        
        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_DIR / f"Corewell_Health_OP_Analysis_{timestamp}.xlsx"
        wb.save(output_file)
        
        logger.info(f"\nExcel report saved to: {output_file}")
        logger.info("=" * 80)
        logger.info("Excel Report Generation Complete!")
        
        return output_file
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        raise

if __name__ == "__main__":
    output_file = main()