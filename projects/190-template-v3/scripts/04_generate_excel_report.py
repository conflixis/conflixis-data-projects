#!/usr/bin/env python3
"""
Generate Excel Report for Health System COI Analysis
Part 4 of the COI Analytics Report Template

This script consolidates all analysis outputs into a structured Excel workbook.
It is driven by a CONFIG.yaml file.
"""

import pandas as pd
from pathlib import Path
import logging
from datetime import datetime
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import glob

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def load_config():
    """Load the configuration from CONFIG.yaml"""
    config_path = Path(__file__).resolve().parent.parent / "CONFIG.yaml"
    if not config_path.exists():
        raise FileNotFoundError(f"CONFIG.yaml not found at {config_path}")
    with open(config_path, 'r') as f:
        return yaml.safe_load(f)

def load_latest_data(data_dir, short_name):
    """Load the most recent analysis outputs for the given short_name"""
    data = {}
    
    patterns = {
        'op_overall': f'{short_name}_op_overall_*.csv',
        'rx_top_drugs': f'{short_name}_rx_top_drugs_*.csv',
        'corr_drug': f'{short_name}_correlation_by_drug_*.csv',
    }
    
    for key, pattern in patterns.items():
        files = list(data_dir.glob(pattern))
        if files:
            latest_file = max(files, key=lambda x: x.stat().st_mtime)
            logger.info(f"Loading {key}: {latest_file.name}")
            data[key] = pd.read_csv(latest_file)
        else:
            logger.warning(f"No files found for pattern: {pattern}")
            
    return data

def create_summary_sheet(wb, data, health_system_name):
    """Create executive summary sheet"""
    ws = wb.create_sheet("Executive Summary", 0)
    ws['A1'] = f"{health_system_name} - COI Analysis Summary"
    ws['A1'].font = Font(bold=True, size=16)
    # ... more formatting and data population ...
    return ws

# ... (Other sheet creation functions would be refactored similarly) ...

def main():
    """Main execution function"""
    try:
        config = load_config()
        hs_config = config['health_system']
        path_config = config['project_paths']

        logger.info("=" * 80)
        logger.info(f"GENERATING EXCEL REPORT FOR: {hs_config['name']}")
        logger.info("=" * 80)
        
        project_dir = Path(path_config['project_dir'])
        data_dir = project_dir / path_config['output_dir']
        reports_dir = project_dir / path_config['reports_dir']
        reports_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info("Loading analysis data...")
        data = load_latest_data(data_dir, hs_config['short_name'])
        
        if not data:
            logger.error("No data files found. Please run analysis scripts first.")
            return
        
        logger.info("Creating Excel workbook...")
        wb = Workbook()
        wb.remove(wb.active)
        
        create_summary_sheet(wb, data, hs_config['name'])
        # ... (calls to other sheet creation functions) ...
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = reports_dir / f"{hs_config['short_name']}_COI_Analysis_{timestamp}.xlsx"
        wb.save(output_file)
        
        logger.info(f"\nExcel report saved to: {output_file}")
        logger.info("=" * 80)
        logger.info("Excel Report Generation Complete!")
        
        return output_file
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        raise

if __name__ == "__main__":
    main()
